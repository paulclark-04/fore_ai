#!/usr/bin/env python3
"""
Fore AI Lead Scoring Agent
Scores LinkedIn profiles for Fore AI SDR outreach (French market focus).
Designed for nuanced inference — not just keyword matching.
"""

import pandas as pd
import re
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import sys
import os
import time

# ═══════════════════════════════════════════════════════════════════════
# REFERENCE LISTS
# ═══════════════════════════════════════════════════════════════════════

ESN_FIRMS = [
    'capgemini', 'sopra steria', 'sopra', 'steria', 'atos', 'cgi',
    'altran', 'alten', 'accenture', 'devoteam', 'aubay', 'tcs',
    'tata consultancy', 'infosys', 'wipro', 'cognizant', 'smile',
    'sqli', 'groupe open', 'sword', 'inetum', 'logica', 'assystem',
    'akka', 'sogeti', 'bearingpoint', 'bearing point', 'wavestone',
    'sia partners', 'mckinsey', 'bain', 'boston consulting', 'bcg',
    'deloitte', 'kpmg', 'ey', 'ernst & young', 'pwc',
    'pricewaterhouse', 'roland berger', 'oliver wyman', 'onepoint',
    'octo technology', 'octo', 'xebia', 'publicis sapient', 'zenika',
    'theodo', 'davidson', 'bemore', 'vision it', 'acial', 'modis',
    'adecco', 'hays', 'michael page', 'randstad', 'manpower',
    'niji', 'thales services', 'econocom', 'open groupe', 'gfi',
]

ELITE_SCHOOLS = [
    'polytechnique', 'x -', "l'x", 'telecom paris', 'télécom paris',
    'centrale', 'mines', 'enseeiht', 'ensimag', 'ensae', 'ensta',
    'insa', 'utbm', 'utc', 'utt', 'epitech', 'epita', 'supelec',
    'isep', 'edhec', 'essec', 'hec ', 'cesi',
]

# ═══════════════════════════════════════════════════════════════════════
# KEYWORD DICTIONARIES
# ═══════════════════════════════════════════════════════════════════════

# --- POSITIVE SIGNALS ---

# Tier 1: Direct QA/Test titles (strongest persona match)
KW_QA_TITLE = [
    'qa', 'quality assurance', 'quality engineering', 'qualité logicielle',
    'software quality', 'test automation', 'test manager', 'testing manager',
    'recette', 'homologation', 'validation logicielle', 'release manager',
    'release management', 'assurance qualité',
]

# Tier 2: DevOps/Pipeline ownership
KW_DEVOPS_TITLE = [
    'devops', 'ci/cd', 'ci cd', 'continuous integration',
    'continuous delivery', 'continuous deployment', 'platform engineering',
    'site reliability', 'sre',
]

# Tier 3: Digital/Product/Transformation (need software context)
KW_DIGITAL_TITLE = [
    'digital', 'numérique', 'transformation', 'product owner',
    'product manager', 'delivery', 'livraison',
]

# Software/Web context words (used to validate ambiguous titles)
KW_SOFTWARE_CONTEXT = [
    'software', 'logiciel', 'web', 'mobile', 'application', 'saas',
    'platform', 'plateforme', 'api', 'frontend', 'backend', 'full stack',
    'fullstack', 'cloud', 'développement', 'development', 'it ',
    'systèmes d\'information', 'informatique', 'digital', 'numérique',
    'microservice', 'architecture logicielle', 'agile', 'scrum',
    'ci/cd', 'devops', 'sprint', 'backlog', 'deploy', 'déploiement',
    'e-commerce', 'e-banking', 'portail', 'portal', 'online',
    'release', 'pipeline', 'testing', 'automation', 'continuous',
    'integration', 'banking', 'banque', 'fintech',
    'data engineering', 'data science', 'data analytics', 'data pipeline',
    'machine learning', 'intelligence artificielle', 'artificial intelligence',
    'mlops', 'dataops', 'analytics platform', 'data platform',
]

# QA automation tools (buying signal — they already use automation)
KW_QA_TOOLS = [
    'selenium', 'cypress', 'playwright', 'hp qtp', 'quicktest',
    'loadrunner', 'testng', 'junit', 'pytest', 'cucumber',
    'robot framework', 'appium', 'katalon', 'testcomplete',
    'ranorex', 'tosca', 'uft', 'testrail', 'xray', 'zephyr',
    'sonarqube', 'mabl', 'testrigor', 'browserstack', 'sauce labs',
    'lambdatest', 'test automation', 'automatisation des tests',
]

# General software tech skills
KW_SOFTWARE_SKILLS = [
    'selenium', 'cypress', 'playwright', 'jira', 'jenkins', 'git',
    'docker', 'kubernetes', 'aws', 'azure', 'gcp', 'java', 'python',
    'javascript', 'typescript', 'react', 'angular', 'vue', 'node',
    'spring', 'j2ee', '.net', 'c#', 'agile', 'scrum', 'kanban',
    'devops', 'ci/cd', 'rest api', 'microservice', 'terraform',
    'ansible', 'puppet', 'sonarqube', 'gitlab', 'github', 'confluence',
    'postman', 'swagger', 'grafana', 'prometheus', 'sql', 'mongodb',
    'postgresql', 'mysql', 'tfs', 'team foundation', 'azure devops',
    'maven', 'gradle', 'bamboo', 'bitbucket',
]

# Buying signal keywords in descriptions/about
KW_BUYING_SIGNALS = [
    'test automation', 'automatisation des tests', 'test strategy',
    'stratégie de test', 'release management', 'gestion des releases',
    'ci/cd', 'continuous', 'regression', 'bottleneck', 'goulot',
    'legacy', 'migration', 'modernization', 'modernisation',
    'transformation digitale', 'digital transformation',
    'user journey', 'parcours utilisateur', 'e2e', 'end-to-end',
    'agile testing', 'shift left', 'test coverage', 'couverture de test',
    'startup', 'poc ', 'proof of concept', 'pilot',
    'scouting', 'veille technologique', 'innovation',
]

# Startup/Innovation signals (can override vertical mismatch)
KW_INNOVATION = [
    'startup', 'scouting', 'partnerships with tech',
    'partenariats avec des startups', 'innovation', 'poc',
    'proof of concept', 'pilot', 'veille technologique',
    'incubat', 'accelerat',
]

# --- NEGATIVE SIGNALS ---

# Non-software QA
KW_RED_NONSOFTWARE = [
    'manufacturing quality', 'qualité industrielle', 'supply chain quality',
    'call center quality', 'qualité centre d\'appels', 'document control',
    'contrôle documentaire', 'qualité produit', 'product quality',
    'qualité fournisseur', 'supplier quality', 'lean manufacturing',
    'six sigma manufacturing', 'qualité process', 'process quality',
    'qualité production', 'production quality', 'quality inspector',
    'contrôle qualité industriel',
]

# Cybersecurity/Fraud
KW_RED_SECURITY = [
    'cybersecurity', 'cybersécurité', 'csirt', 'soc analyst',
    'incident response', 'réponse aux incidents', 'fraud', 'fraude',
    'threat intelligence', 'penetration test', 'ethical hacking',
    'siem', 'intrusion detection', 'malware', 'digital forensics',
    'vulnerability assessment', 'security operations',
]

# Compliance/Audit
KW_RED_COMPLIANCE = [
    'aml analyst', 'kyc quality', 'audit manager', 'risk controls',
    'conformité réglementaire', 'regulatory compliance',
    'anti-money laundering', 'lutte anti-blanchiment',
    'internal audit', 'audit interne', 'contrôle interne',
]

# Physical/Hardware engineering
KW_RED_HARDWARE = [
    'airframe', 'wing equip', 'aile', 'fuselage', 'engine design',
    'moteur', 'manufacturing engineer', 'ingénierie de production',
    'plant management', 'assembly line', 'ligne d\'assemblage',
    'mécanique', 'mechanical engineer', 'hydraulique',
    'tolerancing', 'outillage', 'aérodynamique', 'aerodynamics',
    'structural engineer', 'metallurg', 'tooling engineer',
    'co-working space', 'wing manufacturing',
]

# Titles unambiguously from non-tech domains — skip AI, auto-assign rules score.
# Conservative: only roles where software QA relevance is essentially zero.
# Checked against title_blob (headline + current position title) only.
KW_OBVIOUSLY_IRRELEVANT = [
    # HR / People
    'human resources', 'ressources humaines', 'talent acquisition',
    'talent management', 'recrutement', 'chro', 'chief human',
    'responsable rh', 'hr manager', 'hr director',
    'people & culture', 'people and culture',
    # Marketing (non-tech — "Marketing Technology Manager" is overridden by KW_TECH_MODIFIERS)
    'brand manager', 'brand director', 'directeur de marque',
    'content manager', 'content creator', 'content strategist',
    'copywriter', 'community manager', 'social media manager',
    'social media director', 'marketing manager', 'directeur marketing',
    'chief marketing', 'responsable marketing',
    # Design / Creative (UX/product designers intentionally excluded)
    'graphic designer', 'graphiste', 'art director', 'directeur artistique',
    'creative director', 'directeur créatif', 'motion designer',
    'illustrateur', 'illustrator',
    # Legal
    'general counsel', 'legal director', 'directeur juridique',
    'responsable juridique', 'lawyer', 'avocat', 'notaire', 'juriste',
    'legal counsel', 'chief legal',
    # Finance / Accounting / Treasury
    'chief financial officer', 'directeur financier',
    'financial controller', 'contrôleur de gestion', 'accountant',
    'comptable', 'treasurer', 'trésorier', 'responsable comptable',
    'responsable financier', 'head of finance', 'finance director',
    'chief financial',
    # PR / Communications (non-digital)
    'public relations', 'relations publiques', 'responsable communication',
    'head of communications', 'directeur de la communication',
    'press officer', 'attaché de presse',
    # Facilities / Real Estate
    'facilities manager', 'facility manager', 'gestionnaire immobilier',
    'responsable immobilier', 'head of facilities',
    # Procurement / Purchasing (non-IT)
    'directeur des achats', 'responsable achats',
    'purchasing manager', 'procurement manager',
    # Health / Safety (non-software)
    'médecin', 'infirmier', 'infirmière', 'pharmacien',
    'responsable santé sécurité', 'hse manager',
    'health and safety', 'hygiène sécurité',
]

# Short tokens requiring has_word() to avoid substring matches
# (e.g., 'drh' inside 'sous-drh', 'cmo' inside 'ecommerce')
KW_OBVIOUSLY_IRRELEVANT_WORD = [
    'chro', 'drh', 'cmo', 'cfo', 'clo', 'qhse',
]

# Commercial/CX "digital" role indicators — NOT software engineering or QA
# In retail (e.g., Boulanger), "digital" often means customer experience or e-commerce strategy.
# These titles have no software delivery ownership despite "digital" or "innovation" keywords.
KW_COMMERCIAL_ROLE = [
    'commerce digital', 'directeur commerce', 'responsable commerce',
    'expérience client', 'satisfaction client',
    'expérience et exploitation', 'expérience digitale',
    'directeur offre', 'directrice offre', 'responsable offre',
    'business owner satisfaction', 'business unit director',
]

# Innovation signals that require STRONG tech/startup signals (not just generic "innovation")
# Used to prevent commercial product-offer "innovation" from triggering the innovation scout layer.
KW_INNOVATION_STRONG = [
    'startup', 'scouting', 'partnerships with tech',
    'partenariats avec des startups', 'veille technologique',
    'incubat', 'accelerat',
]

# Revenue/business-focused product role signals (not software QA relevant)
KW_REVENUE_FOCUS = [
    'revenue growth', 'croissance du chiffre', 'conversion rate', 'taux de conversion',
    'monetis', 'monétis', 'p&l', 'ltv', 'arpu', 'acquisition funnel',
    'upsell', 'cross-sell', 'churn reduction', 'b2c monetis', 'b2c revenue',
    'subscription revenue', 'revenu abonnement',
]

# Non-France location signals (French market only)
KW_NON_FRENCH_LOCATIONS = [
    'united states', 'united kingdom', 'germany', 'deutschland',
    'spain', 'españa', 'italy', 'italia', 'netherlands', 'nederland',
    'canada', 'australia', 'india', 'singapore', 'dubai',
    'united arab emirates', ', us', ', usa', ', uk',
    'new york', 'london', 'berlin', 'madrid', 'amsterdam',
    'toronto', 'sydney', 'bangalore', 'mumbai', 'hong kong',
]

# If ANY of these appear in title_blob, override the irrelevant match.
# Intentionally excludes bare 'digital' — "Digital Marketing" is still irrelevant.
KW_TECH_MODIFIERS = [
    'technology', 'technologie', 'informatique',
    "système d'information", "systèmes d'information",
    'information system', 'logiciel', 'software',
    'data', 'cloud', 'platform', 'plateforme',
    'digital transformation', 'transformation digitale',
]

# Seniority keywords by level
KW_SENIORITY_CLEVEL = [
    'cto', 'cio', 'cdo', 'ceo', 'coo', 'chief ',
    'directeur général', 'directrice générale', 'dg ',
    'dsi', 'directeur des systèmes d\'information',
    'general manager', 'président', 'president',
]

KW_SENIORITY_VP = [
    'vp ', 'vice president', 'vice-president', 'v.p.',
    'svp', 'evp', 'senior vice',
]

KW_SENIORITY_DIRECTOR = [
    'directeur', 'directrice', 'director',
]

KW_SENIORITY_HEAD = [
    'head of', 'responsable',
]

KW_SENIORITY_MANAGER = [
    'manager', 'chef de projet', 'chef de programme',
    'project manager', 'program manager', 'team lead', 'tech lead',
    'chapter lead', 'practice manager', 'coordinateur',
    'coordinator', 'lead ',
]


# ═══════════════════════════════════════════════════════════════════════
# UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════

def s(val):
    """Safe lowercase string conversion."""
    if pd.isna(val) or val is None:
        return ''
    return str(val).lower().strip()


def has(text, keywords):
    """Check if text contains any keyword (substring match)."""
    t = s(text)
    return any(kw in t for kw in keywords)


def has_word(text, keywords):
    """Check if text contains any keyword as a whole word (not inside other words).
    Use for short ambiguous keywords like 'cto', 'cio', 'vp', 'dsi'."""
    t = ' ' + s(text) + ' '
    for kw in keywords:
        kw_clean = kw.strip()
        pattern = r'(?<![a-zA-Zà-ÿ])' + re.escape(kw_clean) + r'(?![a-zA-Zà-ÿ])'
        if re.search(pattern, t):
            return True
    return False


def count(text, keywords):
    """Count keyword matches in text."""
    t = s(text)
    return sum(1 for kw in keywords if kw in t)


def get_col(lead, *names):
    """Try multiple column name variants, return first found."""
    for name in names:
        val = lead.get(name)
        if val is not None and not (isinstance(val, float) and pd.isna(val)):
            return str(val)
    return ''


def get_col_from_t(t, field):
    """Get a field from the text dict, with fallback."""
    mapping = {'duration': 'exp0_duration'}
    return t.get(mapping.get(field, field), '')


def extract_years(duration_str):
    """Extract approximate years from duration string like '11 yrs 4 mos'."""
    d = s(duration_str)
    years = 0
    yr_match = re.search(r'(\d+)\s*(?:yr|year|an)', d)
    if yr_match:
        years = int(yr_match.group(1))
    mo_match = re.search(r'(\d+)\s*(?:mo|month|mois)', d)
    if mo_match:
        years += int(mo_match.group(1)) / 12
    return years


def retry_with_backoff(fn, max_retries=3, base_delay=1.0):
    """Retry a function with exponential backoff. Returns result or None."""
    for attempt in range(max_retries):
        result = fn()
        if result is not None:
            return result
        if attempt < max_retries - 1:
            delay = base_delay * (2 ** attempt)
            print(f"    Retrying in {delay}s "
                  f"(attempt {attempt+2}/{max_retries})...",
                  end=' ', flush=True)
            time.sleep(delay)
    return None


# ═══════════════════════════════════════════════════════════════════════
# SIGNAL EXTRACTION FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════

def extract_all_text(lead):
    """Build text blobs at different scopes for analysis."""
    headline = s(get_col(lead, 'headline'))
    about = s(get_col(lead, 'about'))
    current_pos = s(get_col(lead, 'experience/0/position'))
    current_desc = s(get_col(lead, 'experience/0/description'))
    current_company = s(get_col(lead, 'currentPosition/0/companyName',
                                 'experience/0/companyName'))

    # Skills across all experiences
    skills_0 = s(get_col(lead, 'Experience 0 - Skills', 'experience 0 - Skills'))
    skills_1 = s(get_col(lead, 'Experience 1 - Skills', 'experience 1 - Skills'))
    skills_2 = s(get_col(lead, 'experience 2 - Skills', 'Experience 2 - Skills'))
    all_skills = f"{skills_0} {skills_1} {skills_2}"

    # Past positions
    past_pos_1 = s(get_col(lead, 'experience/1/position'))
    past_desc_1 = s(get_col(lead, 'experience/1/description'))
    past_company_1 = s(get_col(lead, 'experience/1/companyName'))
    past_pos_2 = s(get_col(lead, 'experience/2/position'))
    past_desc_2 = s(get_col(lead, 'experience/2/description'))
    past_company_2 = s(get_col(lead, 'experience/2/companyName'))

    exp0_duration = s(get_col(lead, 'experience/0/duration'))
    exp0_location = s(get_col(lead, 'experience/0/location'))
    exp0_employment_type = s(get_col(lead, 'experience/0/employmentType',
                                      'experience/0/employment_type'))
    # The company the lead was *searched for* (from Leads Finder), vs current LinkedIn employer
    searched_company = s(get_col(lead, 'company', 'Company'))

    # Scoped text blobs
    title_blob = f"{headline} {current_pos}"
    current_blob = f"{headline} {current_pos} {current_desc}"
    past_blob = f"{past_pos_1} {past_desc_1} {past_pos_2} {past_desc_2}"
    full_blob = f"{headline} {about} {current_pos} {current_desc} {past_pos_1} {past_desc_1} {past_pos_2} {past_desc_2}"
    all_companies = [current_company, s(past_company_1), s(past_company_2)]

    return {
        'headline': headline, 'about': about,
        'current_pos': current_pos, 'current_desc': current_desc,
        'current_company': current_company,
        'title_blob': title_blob, 'current_blob': current_blob,
        'past_blob': past_blob, 'full_blob': full_blob,
        'all_skills': all_skills, 'all_companies': all_companies,
        'skills_0': skills_0, 'skills_1': skills_1, 'skills_2': skills_2,
        'past_pos_1': past_pos_1, 'past_desc_1': past_desc_1,
        'past_company_1': past_company_1,
        'past_pos_2': past_pos_2, 'past_desc_2': past_desc_2,
        'past_company_2': past_company_2,
        'exp0_duration': exp0_duration,
        'exp0_location': exp0_location,
        'exp0_employment_type': exp0_employment_type,
        'searched_company': searched_company,
    }


# ═══════════════════════════════════════════════════════════════════════
# SCORING COMPONENTS
# ═══════════════════════════════════════════════════════════════════════

def score_persona(t):
    """
    Score persona match (0-40).
    Core question: Does this person own, influence, or feel pain from software QA?
    """
    title = t['title_blob']
    current = t['current_blob']
    full = t['full_blob']
    about = t['about']
    skills = t['all_skills']

    # ── Layer 1: Direct QA/Test title (slam dunk) ──
    if has(title, KW_QA_TITLE):
        # Check it's software QA, not manufacturing QA
        if not has(title, KW_RED_NONSOFTWARE):
            # Career consistency bonus: past roles also QA/testing?
            past_qa = has(t['past_blob'], KW_QA_TITLE + ['testing', 'test ', 'qa ', 'quality assurance'])
            bonus = 5 if past_qa else 0
            return min(40, 35 + bonus), 'Direct QA/Test leadership'

    # ── Layer 1.5: Technical Program/Project Manager with QA automation background ──
    # A TPM who was previously a hands-on test automation engineer understands QA pain
    # at the deepest level — near-primary buyer.
    if has(title, ['technical program manager', 'programme technique',
                    'tpm', 'program manager', 'chef de programme']):
        test_automation_past = (
            has(t['past_blob'], ['test automation', 'automatisation des tests',
                                  'test engineer', 'ingénieur test', 'sdet',
                                  'software development engineer in test',
                                  'qa engineer', 'quality engineer'])
            or has(t['about'], ['automated testing', 'test automation',
                                 'automatisation des tests'])
        )
        if test_automation_past:
            # Include about section — TPMs often describe their work there
            software_ctx = has(t['current_blob'] + ' ' + t['all_skills'] + ' ' + t['about'],
                               KW_SOFTWARE_CONTEXT + ['automated testing', 'test automation',
                                                       'automatisation des tests'])
            if software_ctx:
                return 35, 'Technical Program Manager with QA automation background'

    # ── Layer 2: DevOps/CI-CD pipeline ownership ──
    if has(title, KW_DEVOPS_TITLE):
        # DevOps person owns the pipeline where QA bottlenecks live
        software_ctx = has(current + ' ' + skills, KW_SOFTWARE_CONTEXT)
        # Also check if past roles show digital/platform ownership
        digital_past = has(t['past_blob'], ['digital platform', 'plateforme',
                                             'web', 'application', 'architecture'])
        if software_ctx:
            bonus = 4 if digital_past else 0
            return min(40, 32 + bonus), 'DevOps/CI-CD pipeline owner'
        return 24, 'DevOps (limited software context visible)'

    # ── Layer 3: VP-level Technical Product Owner (check BEFORE C-Level) ──
    # VP + Product Owner/Project Manager in banking = strong buyer
    if has_word(title, KW_SENIORITY_VP + ['vice president', 'vice-president']):
        if has(title, ['product owner', 'product manager', 'project manager',
                        'chef de projet', 'chef de produit']):
            web_context = has(full + ' ' + about, ['web', 'application', 'portal',
                                                     'portail', 'e-commerce',
                                                     'e-banking', 'platform', 'saas',
                                                     'banking', 'banque'])
            technical = has(full + ' ' + skills, ['j2ee', 'java', 'spring',
                                                    '.net', 'python', 'technical',
                                                    'architecture', 'backend',
                                                    'frontend', 'migration', 'legacy'])
            if web_context and technical:
                return 30, 'VP Technical Product Owner (web apps in finance)'
            elif web_context or technical:
                return 25, 'VP Product Owner (technical context)'

    # ── Layer 4: C-Level with Technology/Operations scope ──
    if has_word(title, KW_SENIORITY_CLEVEL):
        # CPO / Chief Product Officer owns product quality at software companies
        # Unlike "Chief Technology Officer", CPO scope is always the product (web app / platform)
        is_cpo = has(title, ['chief product officer']) or has_word(title, ['cpo'])
        if is_cpo and has(current + ' ' + full, KW_SOFTWARE_CONTEXT):
            return 30, 'CPO executive sponsor (product quality owner)'

        # CDO / Chief Data Officer / CDAO — data & AI executives
        # Tightly linked to software quality: data pipelines, data engineering, ML platform.
        # Must be assessed as an executive sponsor, not lumped into "C-Level unclear scope".
        is_cdo = (
            has(title, ['chief data officer', 'chief data & ai', 'chief data and ai',
                        'chief data', 'chief ai officer', 'chief analytics officer',
                        'chief artificial intelligence'])
            or has_word(title, ['cdo', 'cdao'])
        )
        if is_cdo:
            return 28, 'CDO/Chief Data Officer (data & AI executive sponsor)'

        tech_ops = has(current, ['technologies', 'opérations', 'operations',
                                  'systèmes d\'information', 'information system',
                                  'it ', 'numérique', 'technique'])
        # Exclude "digital" when paired with fraud/security
        digital_ok = has(current, ['digital']) and not has(current, ['fraud', 'fraude', 'cyber', 'security'])
        if tech_ops or digital_ok:
            return 30, 'C-Level Tech/Ops executive (executive sponsor)'
        # CTO trap: in manufacturing, CTO = physical product tech
        if has(full, KW_RED_HARDWARE) and not has(full, KW_SOFTWARE_CONTEXT):
            return 5, 'C-Level but hardware/manufacturing focus'
        return 15, 'C-Level (unclear tech scope)'

    # ── Layer 5: Innovation/Startup scout with budget ──
    # Require STRONG innovation signals (startup, scouting, POC, partnerships).
    # Generic "innovation" alone is too common in retail/commercial contexts
    # (e.g., "Directeur de l'innovation de l'Offre" = product category innovation, not tech).
    # Also exclude commercial/CX roles even if they have tech-adjacent descriptions.
    is_commercial_l5 = has(title, KW_COMMERCIAL_ROLE)
    has_strong_innov = has(full, KW_INNOVATION_STRONG)
    # Weak signals (innovation, poc, pilot) only trigger if current title has software context
    has_weak_innov = has(full, ['innovation', 'poc', 'proof of concept', 'pilot'])
    if not is_commercial_l5 and (has_strong_innov or
                                  (has_weak_innov and has(title, KW_SOFTWARE_CONTEXT))):
        budget_signals = has(full, ['budget', '€', 'million', 'roadmap',
                                     'leads a', 'team of', 'équipe'])
        software_past = has(t['past_blob'] + ' ' + skills, KW_SOFTWARE_CONTEXT)
        if budget_signals and software_past:
            return 32, 'Innovation scout with budget + software DNA'
        elif budget_signals:
            return 25, 'Innovation scout with budget'
        return 12, 'Innovation mention (no clear budget)'

    # ── Layer 6: Technical Product Owner/Manager for web apps ──
    if has(title, ['product owner', 'product manager', 'chef de produit', 'head of product']):
        # Exclude CX/UX product roles — "Expérience Digitale", "Self&Care", customer journey
        # Product managers in these domains own customer experience metrics, not software delivery.
        if has(title, KW_COMMERCIAL_ROLE + ['self&care', 'self care', 'parcours client',
                                             'customer journey']):
            return 5, 'CX/UX Product Manager (customer experience, not software delivery)'
        web_context = has(full, ['web', 'application', 'portal', 'portail',
                                  'e-commerce', 'e-banking', 'platform', 'saas'])
        technical = has(full + ' ' + skills, ['j2ee', 'java', 'spring',
                                               '.net', 'python', 'technical',
                                               'architecture', 'backend', 'frontend'])
        if web_context and technical:
            return 25, 'Technical Product Owner (web/digital apps)'
        elif web_context:
            return 18, 'Product Owner (web context)'
        return 8, 'Product Owner (no clear web/software context)'

    # ── Layer 7: Digital Transformation with software context ──
    # Exclude "digital" when paired with fraud/security/compliance
    # Also exclude commercial/CX "digital" roles (Boulanger pattern: "Directeur Commerce Digital",
    # "Directeur Expérience et Exploitation Digitale" = business roles, not software transformation)
    digital_is_valid = has(title, KW_DIGITAL_TITLE) and not has(title,
        ['fraud', 'fraude', 'cyber', 'security', 'sécurité', 'compliance'])
    if digital_is_valid:
        if has(title, KW_COMMERCIAL_ROLE):
            return 5, 'Commercial/CX digital role (not software transformation)'
        software_ctx = has(current + ' ' + about + ' ' + skills,
                           KW_SOFTWARE_CONTEXT)
        if software_ctx:
            return 22, 'Digital/Transformation (software context confirmed)'
        return 10, 'Digital/Transformation (no software context)'

    # ── Layer 8: Engineering/Architecture leadership ──
    if has(title, ['engineering', 'ingénierie', 'architecture', 'architect']):
        if has(current + ' ' + skills, KW_SOFTWARE_CONTEXT):
            if not has(current, KW_RED_HARDWARE):
                return 22, 'Software Engineering/Architecture leadership'
        if has(current, KW_RED_HARDWARE):
            return 0, 'Hardware/Physical engineering'
        return 8, 'Engineering (unclear if software)'

    # ── Layer 8: Project/Delivery management ──
    if has(title, ['project manager', 'chef de projet', 'delivery',
                    'programme manager', 'program manager']):
        if has(full + ' ' + skills, KW_SOFTWARE_CONTEXT):
            return 15, 'IT/Software Project/Delivery management'
        return 5, 'Project management (no software context)'

    # ── Fallback: check if there's software DNA anywhere ──
    if has(full + ' ' + skills, KW_SOFTWARE_CONTEXT):
        return 5, 'Some software context but no clear persona match'

    return 0, 'No persona match detected'


def score_seniority(t):
    """Score seniority level (0-20)."""
    title = t['title_blob']

    # Check VP FIRST (before C-Level) to avoid "vice-president" matching "president"
    if has_word(title, KW_SENIORITY_VP):
        return 18, 'VP'
    # C-Level: use has_word for short acronyms to avoid 'cto' in 'factory' etc
    if has_word(title, KW_SENIORITY_CLEVEL):
        return 20, 'C-Level'
    if has(title, KW_SENIORITY_DIRECTOR):
        return 16, 'Director/Directeur'
    if has(title, KW_SENIORITY_HEAD):
        return 14, 'Head of/Responsable'
    if has(title, KW_SENIORITY_MANAGER):
        # Senior/Staff/Lead/Principal PM or Manager → slight bonus over regular Manager
        if has(title, ['senior', 'principal', 'staff ']):
            return 13, 'Senior Manager/Lead'
        return 10, 'Manager/Lead'
    # Check if senior IC with management scope
    if has(title, ['senior', 'principal', 'staff']):
        if has(t['current_desc'], ['team', 'équipe', 'manage', 'lead',
                                    'encadre', 'supervise']):
            return 8, 'Senior IC with management scope'
        return 4, 'Senior IC'
    return 2, 'IC/Junior/Unclear'


def score_software_dna(t):
    """
    Score software/web DNA across full career (0-25).
    Core question: Is there software engineering in this person's bloodstream?
    """
    score = 0
    signals = []

    # ── Current role software context (highest weight) ──
    current_sw = count(t['current_blob'], KW_SOFTWARE_CONTEXT)
    if current_sw >= 4:
        score += 12
        signals.append('Strong software context in current role')
    elif current_sw >= 2:
        score += 8
        signals.append('Moderate software context in current role')
    elif current_sw >= 1:
        score += 4
        signals.append('Some software context in current role')

    # ── About section software context (supplements current role) ──
    about_sw = count(t['about'], KW_SOFTWARE_CONTEXT)
    if about_sw >= 3:
        score += 6
        signals.append('Rich software context in About section')
    elif about_sw >= 2:
        score += 4
        signals.append('Software context in About section')
    elif about_sw >= 1:
        score += 2

    # ── Past roles software signals (medium weight) ──
    past_sw = count(t['past_blob'], KW_SOFTWARE_CONTEXT)
    if past_sw >= 3:
        score += 6
        signals.append('Software background in past roles')
    elif past_sw >= 1:
        score += 3
        signals.append('Some software in past roles')

    # ── QA automation tools anywhere (strong buying signal) ──
    qa_tools = count(t['all_skills'] + ' ' + t['about'] + ' ' + t['full_blob'],
                     KW_QA_TOOLS)
    if qa_tools >= 3:
        score += 8
        signals.append(f'Heavy QA automation tooling ({qa_tools} tools)')
    elif qa_tools >= 1:
        score += 4
        signals.append('QA automation tool experience')

    # ── General software skills ──
    sw_skills = count(t['all_skills'], KW_SOFTWARE_SKILLS)
    skill_pts = min(sw_skills * 1, 5)
    score += skill_pts
    if sw_skills > 0:
        signals.append(f'{sw_skills} software tech skills')

    return min(score, 30), signals


def score_buying_signals(t):
    """Score buying signals and pain indicators (0-20)."""
    score = 0
    signals = []
    blob = t['full_blob'] + ' ' + t['all_skills'] + ' ' + t['about']

    # Buying signal keywords
    bs_count = count(blob, KW_BUYING_SIGNALS)
    bs_pts = min(bs_count * 3, 12)
    score += bs_pts
    if bs_count > 0:
        signals.append(f'{bs_count} buying signal keywords')

    # Team/department management
    if has(blob, ['team of', 'équipe de', 'leads a team', 'manages a team',
                   'encadre', 'department', 'département', 'chapter',
                   'multifunctional team', 'équipe pluridisciplinaire']):
        score += 3
        signals.append('Manages team/department')

    # Budget authority
    if has(blob, ['budget', '€', 'million', 'm€', 'roadmap',
                   'investment', 'investissement', 'p&l']):
        score += 4
        signals.append('Budget authority signals')

    # Long tenure signal (established authority, budget holder)
    duration = t.get('exp0_duration', '')
    if duration:
        years = extract_years(duration)
        if years >= 8:
            score += 4
            signals.append(f'Long tenure ({years}+ yrs) = established authority')
        elif years >= 4:
            score += 2
            signals.append(f'Solid tenure ({years}+ yrs)')

    return min(score, 20), signals


def score_red_flags(t):
    """Detect red flags (returns penalty 0+, and flag descriptions)."""
    penalty = 0
    flags = []
    current = t['current_blob']
    full = t['full_blob']

    # ── Non-software QA ──
    if has(current, KW_RED_NONSOFTWARE):
        penalty += 30
        flags.append('Non-software QA (manufacturing/supply chain)')

    # ── Cybersecurity/Fraud ──
    # Only check headline + current position title — NOT descriptions
    # (a project mentioning "cybersecurity" shouldn't flag an innovation leader)
    headline_and_title = t['headline'] + ' ' + t['current_pos']
    if has(headline_and_title, KW_RED_SECURITY):
        penalty += 30
        flags.append('Cybersecurity/Fraud operations — not software QA')

    # ── Compliance/Audit ──
    if has(headline_and_title, KW_RED_COMPLIANCE):
        penalty += 25
        flags.append('Compliance/Audit role')

    # ── Hardware/Physical engineering ──
    # Check both title AND descriptions (Cornelius pattern: title is vague
    # but descriptions reveal hardware focus)
    current_with_desc = t['headline'] + ' ' + t['current_pos'] + ' ' + t['current_desc']
    past_with_desc = t['past_blob']
    hardware_current = has(current_with_desc, KW_RED_HARDWARE)
    hardware_past = has(past_with_desc, KW_RED_HARDWARE)
    if hardware_current or hardware_past:
        # Only penalize if no software context overrides it
        has_software = has(current_with_desc + ' ' + t['all_skills'],
                          KW_SOFTWARE_CONTEXT)
        if not has_software:
            hw_penalty = 30 if (hardware_current and hardware_past) else 20
            penalty += hw_penalty
            flags.append('Physical/hardware engineering (no software context)')

    # ── Geography: French market only ──
    exp0_loc = t.get('exp0_location', '')
    if exp0_loc and has(exp0_loc, KW_NON_FRENCH_LOCATIONS):
        penalty += 60
        flags.append(f'Located outside France ({exp0_loc}) — French market only')

    # ── Left target company ──
    # Compare the company we searched for (from Leads Finder) vs current LinkedIn employer.
    # If they differ, this person has moved on — don't target them at the old company.
    searched_co = t.get('searched_company', '')
    current_co = t.get('current_company', '')
    if searched_co and current_co:
        def _norm_co(c):
            return re.sub(r'[^a-z0-9]', '', c.lower())
        ns, nc = _norm_co(searched_co), _norm_co(current_co)
        if ns and nc and ns not in nc and nc not in ns:
            penalty += 50
            flags.append(f'No longer at target company (now at {current_co})')

    # ── Self-employed / coaching signals in headline ──
    # Belt-and-suspenders for when Leads Finder data is stale but headline is updated.
    already_left = any('no longer at target company' in f.lower() for f in flags)
    ex_title_pat = (r'\bex[-\u2013]\s*'
                    r'(head|vp|cto|cpo|coo|cio|director|directeur|'
                    r'manager|lead|chief|responsable)')
    is_ex_title = bool(re.search(ex_title_pat, t['headline']))
    is_helping = bool(re.search(r'\bi help\b', t['headline']))
    if (is_ex_title or is_helping) and not already_left:
        penalty += 40
        flags.append('Self-employed/coaching signals in headline (ex- prefix or coaching framing)')

    # ── Consultant/ESN detection ──
    # Only flag CURRENT employer as ESN. Past ESN is OK if current is different.
    is_esn = False
    esn_name = ''
    current_company = t['current_company']
    for esn in ESN_FIRMS:
        if esn in current_company:
            is_esn = True
            esn_name = current_company
            break
    # Also check current position title for consultant signals
    if not is_esn:
        if has(t['current_pos'], ['consultant', 'freelance', 'prestataire']):
            is_esn = True
            esn_name = 'consultant title in current role'
    if is_esn:
        penalty += 15
        flags.append(f'Likely consultant/ESN ({esn_name})')

    # ── Intern / student detection ──
    emp_type = t.get('exp0_employment_type', '')
    if has(emp_type, ['internship', 'intern', 'stagiaire', 'alternance',
                       'apprentissage', 'work study', 'work-study']):
        penalty += 50
        flags.append('Current role is internship/work-study (no decision authority)')
    # Graduation year in headline: "'25", "25'", "class of 2025", etc.
    grad_year_pat = re.compile(r"['\u2019]2[4-9]|2[4-9]['\u2019]"
                                r"|\bclass of 20(2[4-9]|3[0-2])\b")
    if grad_year_pat.search(t['headline']):
        penalty += 50
        flags.append('Recent graduate/student (graduation year in headline)')

    # ── Zero software DNA across entire profile ──
    if not has(full + ' ' + t['all_skills'] + ' ' + t['about'],
               KW_SOFTWARE_CONTEXT + KW_SOFTWARE_SKILLS):
        penalty += 20
        flags.append('No software/web/digital signals anywhere in profile')

    # ── Business/revenue-focused product role ──
    # A Head/PM of Product focused on revenue growth, CRO, or monetisation owns
    # business metrics — not software quality. Different budget, different pain.
    is_product_role = has(t['title_blob'], ['product owner', 'product manager',
                                             'chef de produit', 'head of product',
                                             'chief product'])
    if is_product_role:
        revenue_count = count(t['full_blob'] + ' ' + t['about'], KW_REVENUE_FOCUS)
        revenue_in_title = has(t['title_blob'] + ' ' + t['headline'],
                               ['revenue', 'growth', 'conversion', 'monetis',
                                'acquisition', 'circularity revenue'])
        if revenue_count >= 2:
            penalty += 15
            flags.append('Product role focused on revenue/conversion (not software QA)')
        elif revenue_count >= 1 and revenue_in_title:
            penalty += 10
            flags.append('Product role with revenue/growth focus (limited QA relevance)')

    # ── Recency rule: past relevant, current not ──
    # If current role has red flags but past roles were relevant, still penalize
    # (current role trumps past history)
    if penalty == 0:
        past_relevant = has(t['past_blob'], KW_QA_TITLE + KW_DEVOPS_TITLE)
        # Check current relevance broadly: include about + skills, not just current_blob
        # (TPMs and technical leaders often describe current work in About section)
        current_irrelevant = not has(current + ' ' + t['about'] + ' ' + t['all_skills'],
                                     KW_SOFTWARE_CONTEXT + KW_QA_TITLE + KW_DEVOPS_TITLE
                                     + ['automated testing', 'test automation', 'program manager',
                                        'technical program', 'delivery', 'release'])
        if past_relevant and current_irrelevant:
            penalty += 10
            flags.append('Moved away from relevant role (current role differs)')

    return penalty, flags


def score_education(t, lead):
    """Small bonus for elite French engineering schools (0-5)."""
    school = s(get_col(lead, 'education/0/schoolName'))
    if has(school, ELITE_SCHOOLS):
        return 4, school
    return 0, ''


# ═══════════════════════════════════════════════════════════════════════
# MAIN SCORING ORCHESTRATOR
# ═══════════════════════════════════════════════════════════════════════

def score_lead(lead):
    """Score a single lead. Returns dict with score, tier, reasoning, etc."""
    t = extract_all_text(lead)

    # Component scores
    persona_pts, persona_label = score_persona(t)
    seniority_pts, seniority_label = score_seniority(t)
    software_pts, software_signals = score_software_dna(t)
    buying_pts, buying_signals = score_buying_signals(t)
    penalty, red_flags = score_red_flags(t)
    school_pts, school_name = score_education(t, lead)

    # Raw score
    raw = persona_pts + seniority_pts + software_pts + buying_pts + school_pts - penalty
    final = max(0, min(100, raw))

    # Tier
    if final >= 80:
        tier = 'A'
    elif final >= 60:
        tier = 'B'
    elif final >= 40:
        tier = 'C'
    else:
        tier = 'D'

    # Special flags
    special = []
    if 'executive sponsor' in persona_label.lower():
        special.append('EXECUTIVE SPONSOR')
    if 'innovation' in persona_label.lower() or 'scout' in persona_label.lower():
        special.append('INNOVATION SCOUT')
    if any('consultant' in f.lower() or 'esn' in f.lower() for f in red_flags):
        special.append('CONSULTANT FLAG')

    # Reasoning
    reasoning = build_reasoning(
        final, tier, persona_label, seniority_label,
        software_signals, buying_signals, red_flags, special, lead, t
    )

    return {
        'score': final,
        'tier': tier,
        'reasoning': reasoning,
        'persona_label': persona_label,
        'seniority': seniority_label,
        'special_flags': ' | '.join(special) if special else '',
        'red_flags_detail': ' | '.join(red_flags) if red_flags else '',
        # Debug breakdown
        '_persona': persona_pts,
        '_seniority': seniority_pts,
        '_software': software_pts,
        '_buying': buying_pts,
        '_school': school_pts,
        '_penalty': penalty,
    }


def build_reasoning(score, tier, persona_label, seniority_label,
                    software_signals, buying_signals, red_flags,
                    special, lead, t):
    """Generate 1-3 sentence reasoning."""
    name = f"{get_col(lead, 'firstName')} {get_col(lead, 'lastName')}".strip()
    company = get_col(lead, 'currentPosition/0/companyName',
                      'experience/0/companyName')
    parts = []

    if tier == 'A':
        if 'Direct QA' in persona_label:
            parts.append(
                f"Textbook primary buyer: direct QA/Test leadership at {company}."
            )
            if buying_signals:
                parts.append(f"Strong buying signals ({', '.join(buying_signals[:2])}).")
            if software_signals:
                parts.append(f"Already invested in automation tooling.")
        elif 'executive sponsor' in persona_label.lower():
            parts.append(
                f"Executive sponsor: {seniority_label} owning technology & operations at {company}."
            )
            parts.append("Pitch transformation/efficiency, not features.")
        elif 'DevOps' in persona_label:
            parts.append(
                f"Owns the CI/CD pipeline at {company} — QA bottlenecks are their direct problem."
            )
        elif 'Innovation' in persona_label or 'scout' in persona_label.lower():
            parts.append(
                f"Innovation/startup scout with budget authority at {company} — pilot opportunity."
            )
            if software_signals:
                parts.append(f"Software DNA: {software_signals[0]}.")
        elif 'Technical Product Owner' in persona_label:
            parts.append(
                f"Technical PO owning web app delivery at {company} — QA speed directly impacts their deadlines."
            )
        else:
            parts.append(
                f"Strong fit: {persona_label} at {company} ({seniority_label})."
            )
            if software_signals:
                parts.append(software_signals[0] + '.')

    elif tier == 'B':
        parts.append(f"{persona_label} at {company} ({seniority_label}).")
        if software_signals:
            parts.append(f"Software DNA: {software_signals[0]}.")
        if red_flags:
            parts.append(f"Watch-out: {red_flags[0]}.")
        elif not software_signals:
            parts.append("Limited visibility into software QA relevance from profile data.")

    elif tier == 'C':
        parts.append(
            f"Borderline: {persona_label} at {company} ({seniority_label})."
        )
        if software_signals:
            parts.append("Some software context but unclear QA connection.")
        if red_flags:
            parts.append(f"Concerns: {red_flags[0]}.")

    else:  # D
        if red_flags:
            primary_flag = red_flags[0]
            parts.append(f"Disqualified: {primary_flag}.")
            if len(red_flags) > 1:
                parts.append(f"Also: {red_flags[1]}.")
        else:
            parts.append(
                f"No relevant persona match. Role ({persona_label}) not connected to software QA."
            )

    return ' '.join(parts)


# ═══════════════════════════════════════════════════════════════════════
# OUTPUT: XLSX WITH FORMATTING
# ═══════════════════════════════════════════════════════════════════════

TIER_COLORS = {
    'A': 'C6EFCE',  # Green
    'B': 'FFEB9C',  # Yellow
    'C': 'FCD5B4',  # Orange
    'D': 'FFC7CE',  # Red
}

TIER_FONT_COLORS = {
    'A': '006100',
    'B': '9C5700',
    'C': '974706',
    'D': '9C0006',
}


def write_output_xlsx(results, output_path):
    """Write scored results to a formatted XLSX file."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Lead Scores'

    # Headers
    headers = [
        'First Name', 'Last Name', 'LinkedIn URL', 'Headline',
        'Company', 'Score', 'Tier', 'Category', 'Method', 'Reasoning',
        'Persona Match', 'Seniority', 'Special Flags', 'Red Flags',
    ]
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor='2F5496')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    body_font = Font(name='Arial', size=10)
    wrap_align = Alignment(vertical='top', wrap_text=True)

    for row_idx, r in enumerate(results, 2):
        method_label = 'AI' if 'ai' in str(r.get('method', '')).lower() else 'Rules'
        category = r.get('category', 'N/A')
        row_data = [
            r['firstName'], r['lastName'], r['linkedinUrl'], r['headline'],
            r['company'], r['score'], r['tier'], category, method_label,
            r['reasoning'], r['persona_label'], r['seniority'],
            r['special_flags'], r['red_flags_detail'],
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.alignment = wrap_align
            cell.border = thin_border

        # Color-code the Tier cell (column 7)
        tier = r['tier']
        tier_cell = ws.cell(row=row_idx, column=7)
        tier_cell.fill = PatternFill('solid', fgColor=TIER_COLORS.get(tier, 'FFFFFF'))
        tier_cell.font = Font(name='Arial', bold=True, size=10,
                              color=TIER_FONT_COLORS.get(tier, '000000'))
        tier_cell.alignment = Alignment(horizontal='center', vertical='top')

        # Color-code the Score cell (column 6)
        score_cell = ws.cell(row=row_idx, column=6)
        score_cell.fill = PatternFill('solid', fgColor=TIER_COLORS.get(tier, 'FFFFFF'))
        score_cell.font = Font(name='Arial', bold=True, size=10,
                              color=TIER_FONT_COLORS.get(tier, '000000'))
        score_cell.alignment = Alignment(horizontal='center', vertical='top')

        # Color-code the Category cell (column 8)
        category_cell = ws.cell(row=row_idx, column=8)
        category_val = r.get('category', 'N/A')
        CATEGORY_COLORS = {
            'Primary Buyer': ('006100', 'C6EFCE'),
            'Sponsor': ('2F5496', 'D6E4F0'),
            'Connector': ('9C5700', 'FFEB9C'),
            'Procurement Gate': ('595959', 'D9D9D9'),
            'Not Relevant': ('9C0006', 'FFC7CE'),
        }
        cat_font_color, cat_fill_color = CATEGORY_COLORS.get(
            category_val, ('000000', 'FFFFFF'))
        category_cell.fill = PatternFill('solid', fgColor=cat_fill_color)
        category_cell.font = Font(name='Arial', size=9, color=cat_font_color)
        category_cell.alignment = Alignment(
            horizontal='center', vertical='top', wrap_text=True)

        # Style the Method cell (column 9)
        method_cell = ws.cell(row=row_idx, column=9)
        method_cell.alignment = Alignment(horizontal='center', vertical='top')
        if 'ai' in str(r.get('method', '')).lower():
            method_cell.font = Font(name='Arial', size=10, color='7030A0')  # Purple for AI
        else:
            method_cell.font = Font(name='Arial', size=10, color='2F5496')  # Blue for rules

    # Column widths
    col_widths = {
        1: 12, 2: 14, 3: 30, 4: 45,
        5: 25, 6: 8, 7: 6, 8: 18, 9: 10, 10: 65,
        11: 30, 12: 18, 13: 22, 14: 45,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze top row
    ws.freeze_panes = 'A2'

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(results) + 1}"

    # ── Summary sheet ──
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Fore AI Lead Scoring Summary'
    ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='2F5496')
    ws2['A3'] = 'Total Leads Scored:'
    ws2['B3'] = len(results)
    ws2['A4'] = 'Tier A (80-100):'
    ws2['B4'] = sum(1 for r in results if r['tier'] == 'A')
    ws2['A5'] = 'Tier B (60-79):'
    ws2['B5'] = sum(1 for r in results if r['tier'] == 'B')
    ws2['A6'] = 'Tier C (40-59):'
    ws2['B6'] = sum(1 for r in results if r['tier'] == 'C')
    ws2['A7'] = 'Tier D (0-39):'
    ws2['B7'] = sum(1 for r in results if r['tier'] == 'D')
    ws2['A9'] = 'Actionable Leads (A+B):'
    ws2['B9'] = sum(1 for r in results if r['tier'] in ('A', 'B'))
    ws2['B9'].font = Font(name='Arial', bold=True, size=12, color='006100')
    # Method stats
    rules_count = sum(1 for r in results if 'ai' not in str(r.get('method', '')).lower())
    ai_count = sum(1 for r in results if 'ai' in str(r.get('method', '')).lower())
    ws2['A11'] = 'Scoring Method:'
    ws2['A11'].font = Font(name='Arial', bold=True, size=10, color='2F5496')
    ws2['A12'] = 'Rules-based:'
    ws2['B12'] = rules_count
    ws2['A13'] = 'AI-reviewed:'
    ws2['B13'] = ai_count

    # Category counts
    ws2['A15'] = 'Lead Categories:'
    ws2['A15'].font = Font(name='Arial', bold=True, size=10, color='2F5496')
    categories = ['Primary Buyer', 'Sponsor', 'Connector',
                  'Procurement Gate', 'Not Relevant', 'N/A']
    for i, cat in enumerate(categories):
        ws2[f'A{16+i}'] = f'  {cat}:'
        ws2[f'B{16+i}'] = sum(
            1 for r in results if r.get('category', 'N/A') == cat)

    for row in range(3, 22):
        ws2.cell(row=row, column=1).font = Font(name='Arial', size=10)
        ws2.cell(row=row, column=2).font = Font(name='Arial', bold=True, size=10)
    ws2['A11'].font = Font(name='Arial', bold=True, size=10, color='2F5496')
    ws2['A15'].font = Font(name='Arial', bold=True, size=10, color='2F5496')
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 12

    # ── Outreach sheet ──
    ws3 = wb.create_sheet('Outreach')
    outreach_headers = [
        'First Name', 'Last Name', 'Title', 'Organization',
        'Email', 'Mobile Phone number', 'Phone number',
        'Linkedin Profile URL', 'Website', 'Country',
    ]
    for col_idx, header in enumerate(outreach_headers, 1):
        cell = ws3.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, r in enumerate(results, 2):
        row_data = [
            r['firstName'],
            r['lastName'],
            r.get('job_title') or r.get('experience/0/position') or '',
            r.get('company') or r.get('currentPosition/0/companyName') or '',
            r.get('email') or '',
            r.get('mobile_number') or '',
            r.get('phone_number') or '',
            r['linkedinUrl'],
            r.get('website') or '',
            r.get('country') or '',
        ]
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.alignment = Alignment(vertical='top')
            cell.border = thin_border

    outreach_col_widths = {1: 14, 2: 16, 3: 35, 4: 25, 5: 30, 6: 18, 7: 18, 8: 40, 9: 30, 10: 12}
    for col, width in outreach_col_widths.items():
        ws3.column_dimensions[get_column_letter(col)].width = width
    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(outreach_headers))}{len(results) + 1}"

    wb.save(output_path)


# ═══════════════════════════════════════════════════════════════════════
# AI INFERENCE (for borderline leads)
# ═══════════════════════════════════════════════════════════════════════

AI_SCORING_PROMPT = """You are a lead scoring agent for Fore AI, a company that sells autonomous QA agents for enterprise web application testing (French market focus).

YOUR TASK: Score this LinkedIn profile on how well they fit as a Fore AI prospect.

SCORING CRITERIA:
1. Does their title/role give them ownership or influence over SOFTWARE QA decisions?
2. Are they in a relevant vertical (Financial Services, Insurance, E-commerce, Travel, Media)?
3. Do they show buying signals (test automation, CI/CD, Selenium, legacy migration, release management)?
4. Do they have decision authority (Manager+, budget holder)?
5. Are there red flags (non-software QA, cybersecurity/fraud, compliance, hardware engineering, consultant/ESN)?

TARGET PERSONAS:
- Primary: QA/Test Leadership (Head of QA, Test Manager, Release Manager)
- Secondary: Tech Leadership (CTO, VP Engineering, DevOps leads — they own the pipeline where QA sits)
- Growing: Digital/Product Leadership (Head of Digital, Product Owners for web apps)
- Niche: Transformation/Delivery (Digital Factory, PMO — if they mention software/IT)
- Special: Innovation/Startup Scouts (if they have budget + mandate to pilot new tech)
- Special: C-Level Executive Sponsors (DG Technologies, DSI — pitch transformation not features)

CRITICAL FRENCH MARKET INSIGHT:
In France, the person overseeing QA often does NOT have "QA" in their title. You must INFER from their full profile whether they have a say in software QA decisions. Ask yourself: "Would I bet $1000 this person would be interested in discussing QA automation agents?"

ANTI-PERSONAS (Disqualify):
- Non-software QA (manufacturing, supply chain, call center quality)
- Cybersecurity/Fraud/CSIRT
- AML/KYC Compliance
- Hardware/physical engineering (airframes, wings, plant operations)
- Consultants at ESN firms (Capgemini, Sopra Steria, Atos, etc.) with no budget authority
- Junior individual contributors

KEY RULES:
- Current role trumps past history
- Recency trumps About section (About may reflect old skills)
- "CTO" in manufacturing/aerospace = physical product technology, NOT software
- "Digital" + "Fraud" = NOT a digital transformation role
- Past ESN employer is OK if current employer is end-client with long tenure
- Innovation/Startup scouts with budget can override vertical mismatches

TRAINING EXAMPLES:
1. Guillaume Tronche — Head of Software QA at Crédit Agricole CIB → Score: 95, Tier A. Textbook primary buyer.
2. Sophie Planchais — Head of Cloud & DevOps at Airbus → Score: 66, Tier B. Owns CI/CD pipeline, QA bottlenecks are her problem.
3. Nicolas Sorre — Head of Smart Factory at Airbus → Score: 81, Tier A. Innovation scout with €6M budget + startup mandate + software/IT DNA.
4. Cornelius Waidelich — Transformation Leader Airbus CTO → Score: 24, Tier D. All hardware engineering (wings, airframes), zero software.
5. Jeremy Signoret — VP Technical PO at Lombard Odier → Score: 80, Tier A. Owns web app delivery in banking, legacy migration pain.
6. Daryouche Khodaï — Cybersecurity & Digital Fraud at BNP Paribas → Score: 0, Tier D. Security ops, not QA.
7. Laurent Benatar — DG Technologies et Opérations at Groupe BPCE → Score: 74, Tier B. Executive sponsor, pitch transformation.

HARD DISQUALIFIERS — Check these FIRST. If any apply, assign Tier D (score ≤ 25) immediately. Do NOT reason around them.
1. OUTSIDE FRANCE: Experience 1 location is outside France (US, UK, Germany, Netherlands, Canada, Australia, India, Singapore, Dubai, etc.) → Tier D, score ≤ 20
2. LEFT TARGET COMPANY: The person's current LinkedIn employer (Experience 1 Company) does not match the company being scored → Tier D. They have no authority there anymore.
3. INTERNSHIP/WORK-STUDY: Experience 1 Employment Type is "Internship", "Contract" immediately following an internship, or "Work Study" → Tier D
4. STUDENT/RECENT GRAD: Headline contains a graduation cohort year like "EDHEC 25'", "HEC'26", "class of 2025" → Tier D (no budget authority)
5. SELF-EMPLOYED COACH: Headline contains "ex-[title]" or "I help [persona] leaders" → they no longer work at the company, no budget → Tier D
6. REVENUE/CRO PRODUCT: The product/PM role is explicitly and primarily about revenue growth, conversion rate optimization, B2C monetisation, P&L, LTV, ARPU — not about software quality or engineering delivery → Tier C or D

SCORING TIERS:
- Tier A (80-100): Direct buyer or strong influence over QA decisions
- Tier B (60-79): Relevant persona, likely interested but not direct buyer
- Tier C (40-59): Unclear fit, some signals but ambiguous
- Tier D (0-39): Wrong persona, wrong domain, or clear red flag

NOW SCORE THIS PROFILE:
{profile_text}

Respond ONLY with valid JSON (no markdown, no backticks):
{"score": <0-100>, "tier": "<A/B/C/D>", "reasoning": "<1-3 sentences>", "persona_label": "<short label>"}"""


def load_full_ai_prompt():
    """Load the full-AI scoring prompt from external file."""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    prompt_path = os.path.join(script_dir, 'scoring_prompt.txt')
    if os.path.exists(prompt_path):
        with open(prompt_path, 'r', encoding='utf-8') as f:
            prompt = f.read()
        print(f"  Loaded prompt from {prompt_path} ({len(prompt)} chars)")
        return prompt
    else:
        print(f"  WARNING: {prompt_path} not found, using built-in AI_SCORING_PROMPT")
        return AI_SCORING_PROMPT


def build_profile_text(lead):
    """Build a readable profile summary for AI scoring."""
    parts = []
    fn = get_col(lead, 'firstName')
    ln = get_col(lead, 'lastName')
    if fn or ln:
        parts.append(f"Name: {fn} {ln}")
    hl = get_col(lead, 'headline')
    if hl:
        parts.append(f"Headline: {hl}")
    about = get_col(lead, 'about')
    if about:
        parts.append(f"About: {about}")
    company = get_col(lead, 'currentPosition/0/companyName', 'experience/0/companyName')
    if company:
        parts.append(f"Current Company: {company}")

    for i in range(3):
        pos = get_col(lead, f'experience/{i}/position')
        comp = get_col(lead, f'experience/{i}/companyName')
        desc = get_col(lead, f'experience/{i}/description')
        dur = get_col(lead, f'experience/{i}/duration')
        loc = get_col(lead, f'experience/{i}/location')
        skills = get_col(lead, f'Experience {i} - Skills', f'experience {i} - Skills')
        if pos or comp:
            exp_parts = [f"\nExperience {i+1}:"]
            if pos:
                exp_parts.append(f"  Position: {pos}")
            if comp:
                exp_parts.append(f"  Company: {comp}")
            if dur:
                exp_parts.append(f"  Duration: {dur}")
            if loc:
                exp_parts.append(f"  Location: {loc}")
            if desc:
                exp_parts.append(f"  Description: {desc}")
            if skills:
                exp_parts.append(f"  Skills: {skills}")
            parts.append('\n'.join(exp_parts))

    school = get_col(lead, 'education/0/schoolName')
    degree = get_col(lead, 'education/0/degree')
    if school:
        edu = f"Education: {school}"
        if degree:
            edu += f" — {degree}"
        parts.append(edu)

    return '\n'.join(parts)


def ai_score_lead(lead, api_key, model='claude-sonnet-4-5-20250929'):
    """Score a lead using Claude API for nuanced inference."""
    try:
        import anthropic
    except ImportError:
        print("ERROR: anthropic package not installed. Run: pip install anthropic --break-system-packages")
        return None

    client = anthropic.Anthropic(api_key=api_key)
    profile_text = build_profile_text(lead)
    prompt = AI_SCORING_PROMPT.format(profile_text=profile_text)

    try:
        response = client.messages.create(
            model=model,
            max_tokens=300,
            messages=[{"role": "user", "content": prompt}],
        )
        text = response.content[0].text.strip()
        # Parse JSON response
        result = json.loads(text)
        return {
            'score': int(result.get('score', 50)),
            'tier': result.get('tier', 'C'),
            'reasoning': result.get('reasoning', ''),
            'persona_label': result.get('persona_label', 'AI-scored'),
        }
    except json.JSONDecodeError:
        # Try to extract JSON from response if wrapped in markdown
        json_match = re.search(r'\{.*\}', text, re.DOTALL)
        if json_match:
            try:
                result = json.loads(json_match.group())
                return {
                    'score': int(result.get('score', 50)),
                    'tier': result.get('tier', 'C'),
                    'reasoning': result.get('reasoning', ''),
                    'persona_label': result.get('persona_label', 'AI-scored'),
                }
            except:
                pass
        print(f"  WARNING: Could not parse AI response for {get_col(lead, 'firstName')} {get_col(lead, 'lastName')}")
        return None
    except Exception as e:
        print(f"  WARNING: AI scoring failed: {e}")
        return None


def _parse_ai_response(text):
    """Parse JSON response from AI, handling markdown wrappers."""
    try:
        result = json.loads(text)
        return {
            'score': int(result.get('score', 50)),
            'tier': result.get('tier', 'C'),
            'category': result.get('category', 'N/A'),
            'reasoning': result.get('reasoning', ''),
            'persona_label': result.get('persona_label', 'AI-scored'),
        }
    except json.JSONDecodeError:
        # Try to extract JSON from markdown-wrapped response
        json_match = re.search(r'\{.*\}', text, re.DOTALL)
        if json_match:
            try:
                result = json.loads(json_match.group())
                return {
                    'score': int(result.get('score', 50)),
                    'tier': result.get('tier', 'C'),
                    'category': result.get('category', 'N/A'),
                    'reasoning': result.get('reasoning', ''),
                    'persona_label': result.get('persona_label', 'AI-scored'),
                        }
            except:
                pass
        return None


def ai_score_lead_full(lead, api_key, model='gemini-2.5-flash',
                        prompt_template=None, provider='google'):
    """Score a lead using AI API with the full-AI prompt.

    Supports providers: 'google' (Gemini) and 'anthropic' (Claude).
    """
    profile_text = build_profile_text(lead)

    if prompt_template is None:
        prompt_template = AI_SCORING_PROMPT

    prompt = prompt_template.replace('{profile_text}', profile_text)

    try:
        if provider == 'google':
            text = _call_gemini(api_key, model, prompt)
        else:
            text = _call_anthropic(api_key, model, prompt)

        if text is None:
            return None

        result = _parse_ai_response(text)
        if result is None:
            name = f"{get_col(lead, 'firstName')} {get_col(lead, 'lastName')}"
            print(f"  WARNING: Could not parse AI response for {name}")
            return None

        # Attach debug info for visualization
        result['_profile_text'] = profile_text
        result['_raw_response'] = text
        return result

    except Exception as e:
        print(f"  WARNING: AI scoring failed: {e}")
        return None


_gemini_client = None


def _get_gemini_client(api_key):
    """Return a lazily-initialized singleton Gemini client."""
    global _gemini_client
    if _gemini_client is None:
        try:
            from google import genai
        except ImportError:
            print("ERROR: google-genai package not installed. Run: pip install google-genai")
            return None
        _gemini_client = genai.Client(api_key=api_key)
    return _gemini_client


def _call_gemini(api_key, model, prompt):
    """Call Google Gemini API and return raw text response."""
    try:
        from google.genai import types
    except ImportError:
        print("ERROR: google-genai package not installed. Run: pip install google-genai")
        return None

    client = _get_gemini_client(api_key)
    if client is None:
        return None
    response = client.models.generate_content(
        model=model,
        contents=prompt,
        config=types.GenerateContentConfig(
            thinking_config=types.ThinkingConfig(thinking_budget=1024)
        ),
    )
    return response.text.strip()


def _call_anthropic(api_key, model, prompt):
    """Call Anthropic Claude API and return raw text response."""
    try:
        import anthropic
    except ImportError:
        print("ERROR: anthropic package not installed. Run: pip install anthropic")
        return None

    client = anthropic.Anthropic(api_key=api_key)
    response = client.messages.create(
        model=model,
        max_tokens=400,
        messages=[{"role": "user", "content": prompt}],
    )
    return response.content[0].text.strip()


# Borderline score thresholds for AI review
AI_REVIEW_LOW = 35   # Deprecated: replaced by is_obviously_irrelevant() — kept for reference
AI_REVIEW_HIGH = 82  # Above this = confident A, skip AI


def is_obviously_irrelevant(t):
    """Return True only if title is unambiguously from a non-tech domain.

    Conservative: if in doubt, return False (let AI score the lead).
    Only checks title_blob — descriptions and About belong to AI reasoning.
    This replaces the AI_REVIEW_LOW numeric gate in hybrid mode.
    """
    title = t['title_blob']

    # Step 1: Check for obviously non-tech role keywords
    irrelevant_match = (
        has(title, KW_OBVIOUSLY_IRRELEVANT)
        or has_word(title, KW_OBVIOUSLY_IRRELEVANT_WORD)
    )
    if not irrelevant_match:
        return False

    # Step 2: Tech modifier override — e.g., "HR Information Systems Director"
    if has(title, KW_TECH_MODIFIERS):
        return False

    return True


# ═══════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════

def process_leads(input_csv, output_xlsx, api_key=None, ai_model='claude-sonnet-4-5-20250929',
                  full_ai=False, provider='google'):
    """Read CSV, score all leads, write formatted XLSX.

    Modes:
    - Rules-only (no api_key): Deterministic scoring only
    - Hybrid (api_key, no full_ai): Rules first, AI for borderline leads
    - Full-AI (api_key + full_ai): Every lead scored by AI
    """
    df = pd.read_csv(input_csv)
    print(f"Loaded {len(df)} leads from {input_csv}")
    print(f"Columns: {list(df.columns)}")

    hybrid_mode = api_key is not None and not full_ai
    full_ai_mode = api_key is not None and full_ai

    if full_ai_mode:
        print(f"\n** FULL-AI MODE: Every lead scored by AI (no deterministic bypass)")
        print(f"   Provider: {provider}  Model: {ai_model}")
    elif hybrid_mode:
        print(f"\n** HYBRID MODE: AI inference enabled for all leads except clearly non-tech titles (rules>{AI_REVIEW_HIGH} skips AI)")
        print(f"   Model: {ai_model}")
    else:
        print("\n** RULES-ONLY MODE: Deterministic scoring (no AI inference)")

    # ═══════════════════════════════════════════════════════════════
    # FULL-AI MODE: Every lead goes through AI
    # ═══════════════════════════════════════════════════════════════
    if full_ai_mode:
        prompt_text = load_full_ai_prompt()
        results = []
        ai_reviewed = 0
        ai_failed = 0

        try:
            for idx, row in df.iterrows():
                lead = row.to_dict()
                name = f"{get_col(lead, 'firstName')} {get_col(lead, 'lastName')}".strip()
                print(f"  [{idx+1}/{len(df)}] {name} -> AI...", end=' ', flush=True)

                ai_result = retry_with_backoff(
                    lambda l=lead: ai_score_lead_full(l, api_key, model=ai_model,
                                                      prompt_template=prompt_text,
                                                      provider=provider),
                    max_retries=3,
                    base_delay=1.0,
                )

                if ai_result:
                    entry = {
                        'firstName': get_col(lead, 'firstName'),
                        'lastName': get_col(lead, 'lastName'),
                        'linkedinUrl': get_col(lead, 'linkedinUrl'),
                        'headline': get_col(lead, 'headline'),
                        'company': get_col(lead, 'currentPosition/0/companyName',
                                           'experience/0/companyName'),
                        'method': 'ai',
                        'score': ai_result['score'],
                        'tier': ai_result['tier'],
                        'category': ai_result.get('category', 'N/A'),
                        'reasoning': ai_result['reasoning'],
                        'persona_label': ai_result.get('persona_label', 'AI-scored'),
                        'seniority': '',
                        'special_flags': '',
                        'red_flags_detail': '',
                    }
                    results.append(entry)
                    ai_reviewed += 1
                    print(f"-> {ai_result['score']}/{ai_result['tier']} ({ai_result.get('category', 'N/A')})")
                else:
                    # Fallback to deterministic scoring
                    result = score_lead(lead)
                    entry = {
                        'firstName': get_col(lead, 'firstName'),
                        'lastName': get_col(lead, 'lastName'),
                        'linkedinUrl': get_col(lead, 'linkedinUrl'),
                        'headline': get_col(lead, 'headline'),
                        'company': get_col(lead, 'currentPosition/0/companyName',
                                           'experience/0/companyName'),
                        'method': 'rules (AI fallback)',
                        'category': 'N/A',
                        **result,
                    }
                    results.append(entry)
                    ai_failed += 1
                    print(f"-> FAILED (fallback rules: {result['score']}/{result['tier']})")

                # Rate limit
                if idx < len(df) - 1:
                    time.sleep(0.5)

        except KeyboardInterrupt:
            print(f"\nInterrupted! Saving {len(results)} partial results...")

        print(f"\nFull-AI pass: {ai_reviewed} scored by AI, {ai_failed} fell back to rules")

    # ═══════════════════════════════════════════════════════════════
    # RULES-ONLY / HYBRID MODE
    # ═══════════════════════════════════════════════════════════════
    else:
        # ── Step 1: Deterministic scoring on ALL leads ──
        results = []
        borderline_indices = []

        for idx, row in df.iterrows():
            lead = row.to_dict()
            result = score_lead(lead)
            entry = {
                'firstName': get_col(lead, 'firstName'),
                'lastName': get_col(lead, 'lastName'),
                'linkedinUrl': get_col(lead, 'linkedinUrl'),
                'headline': get_col(lead, 'headline'),
                'company': get_col(lead, 'currentPosition/0/companyName',
                                   'experience/0/companyName'),
                'method': 'rules',
                'category': 'N/A',
                '_lead': lead,  # Keep raw lead for AI pass
                **result,
            }
            results.append(entry)

            # Flag leads for AI review — skip only if title is obviously non-tech
            if hybrid_mode:
                t_check = extract_all_text(lead)
                if is_obviously_irrelevant(t_check):
                    entry['method'] = 'rules (clearly non-tech title)'
                elif result['score'] <= AI_REVIEW_HIGH:
                    borderline_indices.append(len(results) - 1)

        print(f"\nDeterministic pass complete: {len(results)} leads scored")

        # ── Step 2: AI inference on borderline leads ──
        ai_reviewed = 0
        ai_failed = 0
        if hybrid_mode and borderline_indices:
            print(f"Sending {len(borderline_indices)} borderline leads to AI for deeper analysis...")
            for i, res_idx in enumerate(borderline_indices):
                entry = results[res_idx]
                name = f"{entry['firstName']} {entry['lastName']}"
                rule_score = entry['score']
                rule_tier = entry['tier']
                print(f"  [{i+1}/{len(borderline_indices)}] {name} (rules: {rule_score}/{rule_tier}) -> AI...", end=' ', flush=True)

                ai_result = ai_score_lead(entry['_lead'], api_key, model=ai_model)

                if ai_result:
                    entry['score'] = ai_result['score']
                    entry['tier'] = ai_result['tier']
                    entry['reasoning'] = ai_result['reasoning']
                    entry['persona_label'] = ai_result['persona_label']
                    entry['method'] = 'ai'
                    entry['_rule_score'] = rule_score
                    entry['_rule_tier'] = rule_tier
                    print(f"-> AI: {ai_result['score']}/{ai_result['tier']}")
                    ai_reviewed += 1
                else:
                    print(f"-> FAILED (keeping rules: {rule_score}/{rule_tier})")
                    ai_failed += 1

                if i < len(borderline_indices) - 1:
                    time.sleep(0.5)

            print(f"\nAI pass complete: {ai_reviewed} reviewed, {ai_failed} failed (kept rule score)")
        elif hybrid_mode:
            print("No leads queued for AI -- all scored confidently by rules or flagged as clearly non-tech")

        # Clean up: remove internal _lead reference before output
        for entry in results:
            entry.pop('_lead', None)

    # Sort by score descending
    results.sort(key=lambda x: x['score'], reverse=True)

    write_output_xlsx(results, output_xlsx)

    # Print summary
    tiers = {'A': 0, 'B': 0, 'C': 0, 'D': 0}
    ai_count = sum(1 for r in results if 'ai' in str(r.get('method', '')))
    rules_count = len(results) - ai_count
    for r in results:
        tiers[r['tier']] += 1

    print(f"\n{'='*60}")
    print(f"SCORING COMPLETE -- {len(results)} leads processed")
    print(f"{'='*60}")
    print(f"  Tier A (80-100): {tiers['A']} leads")
    print(f"  Tier B (60-79):  {tiers['B']} leads")
    print(f"  Tier C (40-59):  {tiers['C']} leads")
    print(f"  Tier D (0-39):   {tiers['D']} leads")
    print(f"  -> Actionable (A+B): {tiers['A'] + tiers['B']} leads")
    if hybrid_mode or full_ai_mode:
        print(f"\n  Scoring method:")
        print(f"    Rules-only:  {rules_count} leads")
        print(f"    AI-reviewed: {ai_count} leads")

    # Category summary for full-AI mode
    if full_ai_mode:
        categories = {}
        for r in results:
            cat = r.get('category', 'N/A')
            categories[cat] = categories.get(cat, 0) + 1
        print(f"\n  Categories:")
        for cat, cnt in sorted(categories.items(), key=lambda x: -x[1]):
            print(f"    {cat}: {cnt}")

    print(f"\nOutput saved to: {output_xlsx}")

    return results


def _normalize_name(name):
    """Normalize name for comparison: lowercase, strip accents."""
    import unicodedata
    name = unicodedata.normalize('NFD', name.strip().lower())
    return ''.join(c for c in name if unicodedata.category(c) != 'Mn')


def validate_results(results, ground_truth_path):
    """Compare scored results against ground truth.

    Supports ground_truth.json format:
    - Top-level object with "leads" array
    - Each lead has: name, expected_score, expected_tier, expected_category, score_tolerance
    """
    with open(ground_truth_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Support both flat list and nested {leads: [...]} format
    if isinstance(data, list):
        ground_truth = data
    else:
        ground_truth = data.get('leads', [])

    print(f"\n{'='*60}")
    print(f"VALIDATION AGAINST {ground_truth_path}")
    print(f"{'='*60}")

    passes = 0
    fails = 0
    total = len(ground_truth)

    for gt in ground_truth:
        # Support both "name" and "firstName"/"lastName" formats
        if 'name' in gt:
            gt_name = gt['name']
        else:
            gt_name = f"{gt.get('firstName', '')} {gt.get('lastName', '')}".strip()

        # Match by full name (accent-insensitive)
        match = None
        gt_norm = _normalize_name(gt_name)
        for r in results:
            result_name = f"{r['firstName']} {r['lastName']}".strip()
            if _normalize_name(result_name) == gt_norm:
                match = r
                break

        if not match:
            # Not an error if lead isn't in the input CSV (e.g., ground truth has 19 leads
            # but only 7 in test_leads.csv)
            continue

        expected_tier = gt['expected_tier']
        actual_tier = match['tier']
        expected_score = gt['expected_score']
        actual_score = match['score']
        tolerance = gt.get('score_tolerance', 10)

        tier_ok = actual_tier == expected_tier
        score_ok = abs(actual_score - expected_score) <= tolerance

        expected_cat = gt.get('expected_category')
        actual_cat = match.get('category', 'N/A')
        # Skip category check for rules-only mode (category is N/A)
        cat_ok = (expected_cat is None or actual_cat == 'N/A'
                  or actual_cat == expected_cat)

        if tier_ok and score_ok and cat_ok:
            print(f"  PASS  {gt_name:30s} Score: {actual_score:3d} "
                  f"(expected {expected_score} +/-{tolerance})  "
                  f"Tier: {actual_tier}")
            passes += 1
        else:
            reasons = []
            if not tier_ok:
                reasons.append(
                    f"Tier: got {actual_tier}, expected {expected_tier}")
            if not score_ok:
                reasons.append(
                    f"Score: got {actual_score}, expected {expected_score} "
                    f"+/-{tolerance}")
            if not cat_ok:
                reasons.append(
                    f"Category: got {actual_cat}, expected {expected_cat}")
            print(f"  FAIL  {gt_name:30s} {' | '.join(reasons)}")
            fails += 1

    matched = passes + fails
    print(f"\n  Matched: {matched}/{total} ground truth leads found in results")
    print(f"  Results: {passes}/{matched} passed, {fails}/{matched} failed")
    accuracy = passes / matched * 100 if matched > 0 else 0
    print(f"  Accuracy: {accuracy:.1f}%")
    return passes, fails


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(
        description='Fore AI Lead Scoring Agent — Score LinkedIn profiles for SDR outreach',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Rules-only mode (fast, no API calls)
  python fore_ai_scorer.py leads.csv

  # Full-AI mode with Gemini (recommended)
  python fore_ai_scorer.py leads.csv --full-ai --api-key YOUR_GEMINI_KEY

  # Full-AI mode with Anthropic
  python fore_ai_scorer.py leads.csv --full-ai --api-key sk-ant-xxx --provider anthropic

  # Full-AI with validation
  python fore_ai_scorer.py leads.csv --full-ai --api-key KEY --validate ground_truth.json

  # Hybrid mode (rules + AI for borderline leads)
  python fore_ai_scorer.py leads.csv --hybrid --api-key sk-ant-xxx --provider anthropic
        """,
    )
    parser.add_argument('input_csv', help='Input CSV file with LinkedIn profile data')
    parser.add_argument('-o', '--output', default='scored_leads.xlsx', help='Output XLSX file (default: scored_leads.xlsx)')
    parser.add_argument('--api-key', default=None, help='API key for AI scoring (Gemini or Anthropic)')
    parser.add_argument('--hybrid', action='store_true', help='Enable hybrid mode (rules + AI for borderline)')
    parser.add_argument('--full-ai', action='store_true', help='Full-AI mode: every lead scored by AI')
    parser.add_argument('--provider', default='google', choices=['google', 'anthropic'],
                        help='AI provider (default: google for Gemini)')
    parser.add_argument('--model', default=None,
                        help='AI model override (default: gemini-2.5-flash for Google, claude-haiku-4-5-20251001 for Anthropic)')
    parser.add_argument('--validate', default=None, help='Path to ground_truth.json for score validation')

    args = parser.parse_args()

    # Mutual exclusivity check
    if args.hybrid and args.full_ai:
        print("ERROR: Cannot use --hybrid and --full-ai together. Choose one.")
        sys.exit(1)

    # Resolve API key
    api_key = args.api_key
    if not api_key and (args.hybrid or args.full_ai):
        # Try provider-specific env vars
        if args.provider == 'google':
            api_key = os.environ.get('GOOGLE_API_KEY') or os.environ.get('GEMINI_API_KEY')
        else:
            api_key = os.environ.get('ANTHROPIC_API_KEY')
        if not api_key:
            print("ERROR: --hybrid/--full-ai requires --api-key or appropriate env variable")
            sys.exit(1)

    # Resolve model: use explicit --model, or default based on provider
    model = args.model
    if model is None:
        if args.provider == 'google':
            model = 'gemini-2.5-flash'
        else:
            model = 'claude-haiku-4-5-20251001' if args.full_ai else 'claude-sonnet-4-5-20250929'

    results = process_leads(args.input_csv, args.output, api_key=api_key,
                           ai_model=model, full_ai=args.full_ai,
                           provider=args.provider)

    if args.validate:
        validate_results(results, args.validate)
